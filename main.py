# =========================================================
# 환경오염 해결 아이디어 + 생성형 AI 대화 앱 (Streamlit)
# - Solar API(모델: solar-open2)를 openai 라이브러리로 사용
# - 초보자를 위해 한국어 주석을 최대한 자세히 달았어요.
# =========================================================

# 1) 필요한 라이브러리 불러오기 ----------------------------
import io                                     # 엑셀 파일을 메모리에서 만들 때 사용
import time                                   # 타이머(경과 시간 계산)에 사용
import datetime                               # 활동지에 오늘 날짜를 넣을 때 사용
import streamlit as st                        # 화면(웹앱)을 만들어 주는 라이브러리
import streamlit.components.v1 as components  # 실시간 타이머(HTML/JS) 넣을 때 사용
from openai import OpenAI                     # Solar API를 호출할 때 쓰는 라이브러리


# 2) 페이지 기본 설정 --------------------------------------
st.set_page_config(page_title="환경오염 해결 아이디어", page_icon="🌍")


# 3) AI에게 줄 '성격' (시스템 프롬프트) --------------------
SYSTEM_PROMPT = "너는 따뜻하고 친절한 데이터 분석 선생님이야. 반드시 순수 한국어로만 답해"


# 4) Solar API 연결 클라이언트 만들기 ----------------------
client = OpenAI(
    api_key=st.secrets["SOLAR_API_KEY"],   # 코드에 쓰지 않고 비밀 금고에서 불러옴
    base_url="https://api.upstage.ai/v1",  # Solar API 접속 주소
)


# 5) 세션 상태(기억 장치) 준비하기 -------------------------
if "messages" not in st.session_state:      # 진행 중인 대화 (AI가 이전 대화를 기억)
    st.session_state.messages = []
if "records" not in st.session_state:       # 이름별 저장 기록 (이름이 '코드' 역할)
    st.session_state.records = {}
if "timer_start" not in st.session_state:   # 타이머를 시작한 시각 (없으면 None)
    st.session_state.timer_start = None


# 6) '현재 작업 초기화' 실제 처리 --------------------------
# 아래 초기화 버튼을 누르면 이 부분이 먼저 실행되어
# 학생 정보 / 아이디어 입력칸 / AI 대화가 지워집니다. (저장된 기록은 그대로 유지)
if st.session_state.get("_현재작업초기화"):
    st.session_state["_현재작업초기화"] = False
    st.session_state.messages = []               # AI 대화 내용 지우기
    st.session_state.timer_start = None          # 타이머도 초기화
    for 키 in ("이름", "학년선택", "아이디어입력"):  # 입력칸 값들 지우기
        st.session_state.pop(키, None)


# 7) 사이드바(왼쪽 메뉴) - 이름 / 학년 / 초기화 / 암호 -----
with st.sidebar:
    st.header("학생 정보")

    # 이름: 이 이름이 '하나의 코드(구분자)'가 되어
    #      아이디어/프롬프트/결과물을 이 이름에 묶어 저장합니다.
    name = st.text_input("이름을 입력하세요", key="이름")

    # 나이(학년): 초등학교 5학년 / 6학년 중에서 선택
    grade = st.radio("나이(학년)를 선택하세요", ["초등학교 5학년", "초등학교 6학년"], key="학년선택")

    # 현재 작업 초기화 버튼 (저장된 기록은 남고, 지금 쓰던 것만 지움)
    if st.button("🧹 현재 작업 초기화 (기록은 유지)"):
        st.session_state["_현재작업초기화"] = True
        st.rerun()  # 화면을 새로 그려서 깨끗하게 비웁니다.

    st.divider()
    # 기록 확인 암호: '4087'을 맞게 입력한 사람만 저장된 기록을 볼 수 있어요.
    암호 = st.text_input("기록 확인 암호", type="password", placeholder="연구자만 입력")
    암호맞음 = (암호 == "4087")


# 8) 화면 맨 위 제목 - 시험지(활동지)처럼, 제목은 정중앙 두 줄
오늘 = datetime.date.today().strftime("%Y년 %m월 %d일")
이름_표시 = name if name else "&nbsp;" * 10  # 이름이 없으면 빈 칸처럼 보이게

st.markdown(
    f"""
    <div style="border:2px solid #2b2b2b; border-radius:10px; padding:18px 22px;
                background:#fafafa; margin-bottom:14px;">
      <div style="display:flex; justify-content:space-between; align-items:center;
                  font-size:13px; color:#555; letter-spacing:1px;">
        <span>🌱 환경 · 데이터 수업 활동지</span>
        <span>{오늘}</span>
      </div>
      <div style="border-top:1px dashed #bbb; margin:10px 0;"></div>
      <!-- 제목: 정중앙 정렬 + 매끄럽게 두 줄 (word-break:keep-all 로 단어가 잘리지 않게) -->
      <div style="text-align:center; font-size:27px; font-weight:800; color:#1a1a1a;
                  line-height:1.55; word-break:keep-all;">
        환경오염을 해결하기 위한<br>아이디어를 작성해 주세요
      </div>
      <div style="border-top:1px dashed #bbb; margin:12px 0 10px 0;"></div>
      <div style="display:flex; gap:26px; font-size:15px; color:#333;">
        <span>이름 : <b>{이름_표시}</b></span>
        <span>학년 : <b>{grade}</b></span>
        <span>점수 : ______</span>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)


# 9) 상단 타이머 (10분 카운트다운) -------------------------
TIMER_초 = 600  # 10분 = 600초

버튼1, 버튼2, _ = st.columns([1, 1, 3])
with 버튼1:
    if st.button("⏱ 시작"):
        st.session_state.timer_start = time.time()  # 지금 시각 기록 -> 카운트다운 시작
with 버튼2:
    if st.button("리셋"):
        st.session_state.timer_start = None          # 다시 10:00 으로

if st.session_state.timer_start is not None:
    흐른시간 = time.time() - st.session_state.timer_start
    남은초 = max(0, int(round(TIMER_초 - 흐른시간)))
    실행중 = 남은초 > 0
else:
    남은초 = TIMER_초
    실행중 = False

# 실제 화면 시계는 브라우저(JS)가 1초마다 스스로 줄여 보여 줍니다.
# -> 채팅하는 동안에도 앱이 멈추지 않고 시계가 계속 흐릅니다.
타이머_HTML = f"""
<div style="font-family:-apple-system,'Segoe UI',sans-serif; text-align:center;">
  <div id="disp" style="font-size:46px; font-weight:800; letter-spacing:2px; color:#1a1a1a;">10:00</div>
  <div id="msg" style="font-size:13px; color:#888; margin-top:2px;">남은 시간 (10분)</div>
</div>
<script>
  let remaining = {남은초};
  const running = {"true" if 실행중 else "false"};
  const disp = document.getElementById('disp');
  const msg  = document.getElementById('msg');
  function fmt(s){{
    const m = Math.floor(s/60), c = s%60;
    return String(m).padStart(2,'0') + ':' + String(c).padStart(2,'0');
  }}
  function render(){{
    disp.textContent = fmt(remaining);
    if(remaining<=0){{ disp.style.color='#c0392b'; msg.textContent='시간 종료!'; }}
  }}
  render();
  if(running){{
    const id = setInterval(function(){{
      remaining -= 1;
      if(remaining<=0){{ remaining=0; render(); clearInterval(id); }}
      else {{ render(); }}
    }}, 1000);
  }}
</script>
"""
components.html(타이머_HTML, height=110)


# 10) 이름별 저장 공간을 만들어 주는 도우미 함수 -----------
def 이름_저장공간_준비(학생이름):
    """해당 이름의 저장 공간이 없으면 새로 만들어 줍니다."""
    if 학생이름 not in st.session_state.records:
        st.session_state.records[학생이름] = {
            "학년": grade,   # 선택한 학년
            "아이디어": [],   # 학생이 제출한 아이디어들
            "대화": [],      # (프롬프트, 결과물) 쌍으로 저장되는 AI 대화
        }
    st.session_state.records[학생이름]["학년"] = grade  # 항상 최신 학년으로 갱신


# =========================================================
# 11) [첫 번째 칸] 아이디어 작성 칸
# =========================================================
st.header("✏️ 아이디어 작성 칸")
st.caption("환경오염을 해결할 나만의 아이디어를 자유롭게 적어 보세요.")

아이디어_내용 = st.text_area(
    "나의 아이디어", height=140,
    placeholder="예) 학교에서 분리수거 게임을 만들어요!", key="아이디어입력",
)

if st.button("아이디어 제출하기"):
    if not name:
        st.warning("먼저 사이드바에서 이름을 입력해 주세요.")
    elif not 아이디어_내용.strip():
        st.warning("아이디어 내용을 입력해 주세요.")
    else:
        이름_저장공간_준비(name)
        st.session_state.records[name]["아이디어"].append(아이디어_내용)
        st.success(f"'{name}' 학생의 아이디어가 제출되었어요! 👍")

# 제출한 아이디어는 '암호를 맞게 입력한 사람'만 볼 수 있어요.
if name and name in st.session_state.records and st.session_state.records[name]["아이디어"]:
    if 암호맞음:
        제출된 = st.session_state.records[name]["아이디어"]
        with st.expander(f"📌 '{name}' 학생이 제출한 아이디어 보기 ({len(제출된)}개)"):
            for 번호, 글 in enumerate(제출된, start=1):
                st.write(f"{번호}. {글}")
    else:
        st.info("🔒 제출한 아이디어는 사이드바에 암호를 입력해야 볼 수 있어요.")


st.divider()  # 두 칸을 구분하는 선


# =========================================================
# 12) [두 번째 칸] 생성형 AI 대화 칸
# =========================================================
st.header("🤖 생성형 AI 대화 칸")
st.caption("아이디어에 대해 AI 선생님과 이야기해 보세요.")

# 대화가 길어져도 위쪽 '아이디어 작성 칸'이 사라지지 않도록,
# 대화 내용은 높이가 고정된 '스크롤 상자' 안에서만 흐르게 만듭니다.
대화상자 = st.container(height=420)

with 대화상자:
    for 메시지 in st.session_state.messages:
        with st.chat_message(메시지["role"]):
            st.markdown(메시지["content"])


def 실시간_글자흐름(스트림):
    """AI가 보내주는 조각(chunk)에서 글자만 뽑아 하나씩 내보냅니다."""
    for 조각 in 스트림:
        내용 = 조각.choices[0].delta.content
        if 내용:
            yield 내용


질문 = st.chat_input("메시지를 입력하세요")

if 질문:
    if not name:
        st.warning("사이드바에서 이름을 먼저 입력해 주세요. 이름으로 대화가 저장됩니다.")
        st.stop()

    with 대화상자:
        st.session_state.messages.append({"role": "user", "content": 질문})
        with st.chat_message("user"):
            st.markdown(질문)

        with st.chat_message("assistant"):
            try:
                스트림 = client.chat.completions.create(
                    model="solar-open2",       # 모델 이름은 반드시 이 글자 그대로!
                    messages=(
                        [{"role": "system", "content": SYSTEM_PROMPT}]
                        + st.session_state.messages
                    ),
                    stream=True,               # 스트리밍(글자가 실시간으로 흐름)
                    reasoning_effort="none",   # 추론(생각) 끄기 -> 답이 더 빨리 나옴
                )
                답변 = st.write_stream(실시간_글자흐름(스트림))
            except Exception:
                답변 = None
                st.error(
                    "앗, 답변을 가져오는 데 문제가 생겼어요. 😢\n\n"
                    "잠시 후 다시 시도해 주세요. 문제가 계속되면 인터넷 연결이나 "
                    "API 키(SOLAR_API_KEY) 설정을 확인해 주세요."
                )

    if 답변:
        st.session_state.messages.append({"role": "assistant", "content": 답변})
        이름_저장공간_준비(name)
        st.session_state.records[name]["대화"].append(
            {"프롬프트": 질문, "결과물": 답변}
        )


# =========================================================
# 13) 기록을 보기 좋은 Excel(xlsx)로 만들어 주는 함수
#     - 학생별로 시트를 나누고, [제출한 아이디어] -> [AI 대화] 순서로 정리
# =========================================================
def _시트이름(원본, 사용중):
    """엑셀 시트 이름에 쓸 수 없는 글자를 지우고, 31자 제한/중복을 처리합니다."""
    금지 = '[]:*?/\\'
    이름 = "".join(c for c in 원본 if c not in 금지).strip() or "학생"
    이름 = 이름[:28]
    후보, n = 이름, 1
    while 후보 in 사용중:
        n += 1
        후보 = f"{이름}_{n}"[:31]
    사용중.add(후보)
    return 후보


def 기록_excel만들기(대상기록):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    선 = Side(style="thin", color="D9D9D9")
    테두리 = Border(left=선, right=선, top=선, bottom=선)
    가운데 = Alignment(horizontal="center", vertical="center", wrap_text=True)
    왼쪽 = Alignment(horizontal="left", vertical="top", wrap_text=True)
    제목폰트 = Font(bold=True, size=14, color="FFFFFF")
    제목채움 = PatternFill("solid", fgColor="2E7D32")
    구획폰트 = Font(bold=True, size=12, color="1B5E20")
    구획채움 = PatternFill("solid", fgColor="E8F5E9")
    헤더폰트 = Font(bold=True, color="FFFFFF")
    헤더채움 = PatternFill("solid", fgColor="66BB6A")
    라벨폰트 = Font(bold=True)

    사용중 = set()
    for 이름, 정보 in 대상기록.items():
        ws = wb.create_sheet(title=_시트이름(이름, 사용중))
        ws.column_dimensions["A"].width = 8    # 번호
        ws.column_dimensions["B"].width = 50   # 아이디어 / 학생 프롬프트
        ws.column_dimensions["C"].width = 55   # AI 답변

        r = 1
        # 큰 제목
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(r, 1, "환경 수업 활동 기록")
        c.font, c.fill, c.alignment = 제목폰트, 제목채움, 가운데
        ws.row_dimensions[r].height = 26
        r += 1

        # 이름 / 학년
        for 라벨, 값 in [("이름", 이름), ("학년", 정보["학년"])]:
            c = ws.cell(r, 1, 라벨); c.font, c.alignment, c.border = 라벨폰트, 가운데, 테두리
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            ws.cell(r, 2, 값).alignment = 왼쪽
            ws.cell(r, 2).border = 테두리
            ws.cell(r, 3).border = 테두리
            r += 1
        r += 1

        # ── 구획 1: 제출한 아이디어 ──
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(r, 1, "■ 제출한 아이디어"); c.font, c.fill, c.alignment = 구획폰트, 구획채움, 왼쪽
        r += 1
        ws.cell(r, 1, "번호")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(r, 2, "아이디어 내용")
        for col in (1, 2, 3):
            cc = ws.cell(r, col); cc.font, cc.fill, cc.alignment, cc.border = 헤더폰트, 헤더채움, 가운데, 테두리
        r += 1
        if 정보["아이디어"]:
            for i, 글 in enumerate(정보["아이디어"], start=1):
                ws.cell(r, 1, i).alignment = 가운데; ws.cell(r, 1).border = 테두리
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
                ws.cell(r, 2, 글).alignment = 왼쪽
                ws.cell(r, 2).border = 테두리; ws.cell(r, 3).border = 테두리
                r += 1
        else:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.cell(r, 1, "(제출한 아이디어가 없습니다)").alignment = 왼쪽
            r += 1
        r += 1

        # ── 구획 2: 생성형 AI 대화 기록 ──
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(r, 1, "■ 생성형 AI 대화 기록"); c.font, c.fill, c.alignment = 구획폰트, 구획채움, 왼쪽
        r += 1
        for col, txt in [(1, "번호"), (2, "학생 프롬프트"), (3, "AI 답변")]:
            cc = ws.cell(r, col, txt); cc.font, cc.fill, cc.alignment, cc.border = 헤더폰트, 헤더채움, 가운데, 테두리
        r += 1
        if 정보["대화"]:
            for i, 쌍 in enumerate(정보["대화"], start=1):
                ws.cell(r, 1, i).alignment = 가운데; ws.cell(r, 1).border = 테두리
                ws.cell(r, 2, 쌍["프롬프트"]).alignment = 왼쪽; ws.cell(r, 2).border = 테두리
                ws.cell(r, 3, 쌍["결과물"]).alignment = 왼쪽; ws.cell(r, 3).border = 테두리
                r += 1
        else:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.cell(r, 1, "(대화 기록이 없습니다)").alignment = 왼쪽
            r += 1

    저장통 = io.BytesIO()
    wb.save(저장통)
    return 저장통.getvalue()


# =========================================================
# 14) 사이드바 아래쪽 - 암호를 맞힌 연구자만 기록 관리
# =========================================================
with st.sidebar:
    st.divider()
    st.subheader("저장된 기록 (연구자용)")

    if not 암호맞음:
        st.caption("🔒 기록을 보려면 위의 '기록 확인 암호'를 바르게 입력하세요.")
    elif not st.session_state.records:
        st.caption("아직 저장된 기록이 없어요.")
    else:
        # 어떤 학생 기록을 내려받을지 선택 ('(전체)'면 모든 학생)
        이름목록 = ["(전체)"] + list(st.session_state.records.keys())
        선택 = st.selectbox("학생 선택", 이름목록)

        if 선택 == "(전체)":
            대상 = st.session_state.records
            파일이름 = "환경수업_전체기록"
        else:
            대상 = {선택: st.session_state.records[선택]}
            파일이름 = f"{선택}_기록"

        # Excel(xlsx) 로 내려받기
        st.download_button(
            label="📊 Excel 로 내려받기",
            data=기록_excel만들기(대상),
            file_name=f"{파일이름}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.divider()
        # 저장된 기록 전체 삭제 (암호를 맞힌 연구자만 볼 수 있는 버튼)
        st.caption("⚠️ 아래는 저장된 모든 기록을 완전히 지웁니다.")
        확인 = st.checkbox("삭제에 동의합니다")
        if st.button("🗑 저장된 기록 전체 삭제", disabled=not 확인):
            st.session_state.records = {}   # 모든 저장 기록 비우기
            st.rerun()
