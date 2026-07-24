# =========================================================
# common.py : 두 페이지(인지과제 / 창의적 글쓰기)가 함께 쓰는 공통 기능 모음
# - 학생 정보(이름·학년)를 한 번만 입력해도 모든 페이지에서 이어지게 합니다.
# - 10분 타이머, 기록 저장, 기록 내려받기 기능도 여기에 모아 두었어요.
# =========================================================

import io
import csv
import time
import datetime
import streamlit as st
from openai import OpenAI
import supabase_db          # Supabase(온라인 DB) 저장 기능


# ---------------------------------------------------------
# 1) AI에게 줄 '성격' (시스템 프롬프트)
#    - 정답은 절대 알려주지 않고, 스스로 생각하도록 도와주는(비계) 역할입니다.
# ---------------------------------------------------------
SYSTEM_PROMPT = (
    "너는 따뜻하고 친절한 데이터 분석 선생님이야. 반드시 순수 한국어로만 답해. "
    "정답을 알려달라고 해도 절대 정답을 알려주지 마. 어떤 일이 있어도 정답을 이야기하면 안돼. "
    "대신 학생이 스스로 답을 찾도록 도와줘. 문제를 더 작은 단계로 쪼개 주거나, "
    "무엇부터 생각하면 좋을지 되묻거나, 비슷하지만 더 쉬운 예를 들어 주는 방식으로 도와줘. "
    "최종 답이나 최종 계산 결과는 절대 말하지 말고, 초등학생이 이해할 수 있는 쉬운 말로 설명해."
)


# ---------------------------------------------------------
# 2) Solar API 연결 클라이언트 만들기
# ---------------------------------------------------------
def 클라이언트():
    """Solar API에 연결하는 통로를 만들어 돌려줍니다."""
    return OpenAI(
        api_key=st.secrets["SOLAR_API_KEY"],   # 코드에 쓰지 않고 비밀 금고에서 불러옴
        base_url="https://api.upstage.ai/v1",  # Solar API 접속 주소
    )


# ---------------------------------------------------------
# 3) 세션 상태(기억 장치) 준비하기
#    - 세션 상태는 페이지를 옮겨 다녀도 값이 그대로 남는 '기억 장치'예요.
# ---------------------------------------------------------
def 세션_준비():
    if "_이름" not in st.session_state:
        st.session_state["_이름"] = ""                    # 학생 이름 (페이지 공통)
    if "_학년" not in st.session_state:
        st.session_state["_학년"] = "초등학교 5학년"        # 학년 (페이지 공통)
    if "records" not in st.session_state:
        st.session_state["records"] = {}                  # 이름별 저장 기록
    if "_ver" not in st.session_state:
        st.session_state["_ver"] = 0                      # 입력칸 '버전' (초기화용)


# ---------------------------------------------------------
# 4) 사이드바에 학생 정보(이름·학년)를 그려 주는 함수
#    - 한 페이지에서 입력하면 다른 페이지에도 그대로 이어집니다.
#    - 비밀은 '_이름', '_학년'이라는 저장용 칸에 값을 옮겨 두는 것이에요.
#      (위젯에 붙은 값은 페이지를 옮기면 사라질 수 있어서 따로 보관합니다.)
# ---------------------------------------------------------
def 학생정보_사이드바(페이지키):
    학년목록 = ["초등학교 5학년", "초등학교 6학년"]
    버전 = st.session_state.get("_ver", 0)   # 초기화할 때마다 1씩 올라가는 번호

    with st.sidebar:
        st.header("학생 정보")

        # 이름: 이 이름이 '하나의 코드(구분자)'가 되어
        #      정답과 AI 프롬프트/결과물이 이 이름에 묶여 저장됩니다.
        이름 = st.text_input(
            "이름을 입력하세요",
            value=st.session_state["_이름"],       # 이전에 입력한 값을 그대로 보여 줌
            key=f"이름칸_{페이지키}_{버전}",
        )

        # 나이(학년): 초등학교 5학년 / 6학년 중에서 선택
        학년 = st.radio(
            "나이(학년)를 선택하세요",
            학년목록,
            index=학년목록.index(st.session_state["_학년"]),
            key=f"학년칸_{페이지키}_{버전}",
        )

        # 입력한 값을 공통 보관함에 저장 -> 다른 페이지에서도 그대로 사용
        st.session_state["_이름"] = 이름
        st.session_state["_학년"] = 학년

        if 이름:
            st.success(f"👤 {이름} · {학년}")
        else:
            st.info("이름을 입력하면 모든 페이지에서 함께 사용돼요.")

    return 이름, 학년


# ---------------------------------------------------------
# 5) 10분 타이머
#    - 페이지마다 따로 돌아갑니다. (인지과제 10분, 글쓰기 10분)
#    - 시간이 끝나면 True(종료)를 돌려주어 입력칸을 잠그게 합니다.
# ---------------------------------------------------------
def _타이머상태(페이지키):
    """페이지마다 '지금까지 흐른 시간(경과)'과 '켜진 시각(시작)'을 따로 보관합니다."""
    st.session_state.setdefault("_타이머", {})
    return st.session_state["_타이머"].setdefault(페이지키, {"경과": 0.0, "시작": None})


def _다른페이지_타이머_멈춤(현재페이지키):
    """다른 페이지의 타이머를 멈춥니다.

    스트림릿은 페이지를 옮기면 그 페이지의 코드가 실행돼요.
    그때 '지금 보고 있는 페이지'가 아닌 타이머는 모두 멈춰서,
    보고 있는 페이지의 시계만 흐르게 만듭니다.
    """
    지금 = time.time()
    for 키, 상태 in st.session_state.get("_타이머", {}).items():
        if 키 != 현재페이지키 and 상태["시작"] is not None:
            상태["경과"] += 지금 - 상태["시작"]   # 흐른 만큼 쌓아 두고
            상태["시작"] = None                  # 시계를 멈춥니다.


def _남은초계산(페이지키, 총초):
    """지금 남은 시간(초)을 계산합니다."""
    상태 = _타이머상태(페이지키)
    경과 = 상태["경과"]
    if 상태["시작"] is not None:
        경과 += time.time() - 상태["시작"]
    return max(0, int(round(총초 - 경과)))


def _시간글자(초):
    """초를 08:35 같은 모양으로 바꿉니다."""
    return f"{초 // 60:02d}:{초 % 60:02d}"


def 타이머(페이지키, 분=10):
    """10분 타이머. (남은초, 종료여부)를 돌려줍니다.

    - '⏱ 시작'을 눌러야 시계가 흐르기 시작합니다.
    - 다른 페이지로 넘어가는 순간 이 페이지 시계는 자동으로 멈춥니다.
      (돌아와서 '⏱ 시작'을 다시 누르면 멈췄던 지점부터 이어집니다)
    - 시간이 끝나면 종료=True 가 되어 입력칸이 잠깁니다.
    """
    총초 = 분 * 60
    _다른페이지_타이머_멈춤(페이지키)      # 다른 페이지 시계는 멈춤
    상태 = _타이머상태(페이지키)

    버튼1, 가운데, 버튼2 = st.columns([1, 2, 1])
    with 버튼1:
        # 학생이 직접 눌러야 시계가 흐르기 시작합니다.
        if st.button("⏱ 시작", key=f"시작_{페이지키}", use_container_width=True):
            if 상태["시작"] is None and _남은초계산(페이지키, 총초) > 0:
                상태["시작"] = time.time()
            st.rerun()
    with 버튼2:
        if st.button("리셋", key=f"리셋_{페이지키}", use_container_width=True):
            상태["경과"], 상태["시작"] = 0.0, None   # 다시 10:00 (멈춘 상태)
            st.rerun()

    남은초 = _남은초계산(페이지키, 총초)
    실행중 = 상태["시작"] is not None
    시작한적있음 = 실행중 or 상태["경과"] > 0

    # 시간이 다 됐으면 시계를 멈춰 둡니다.
    if 남은초 <= 0 and 실행중:
        상태["시작"] = None
        상태["경과"] = 총초
        실행중 = False

    종료 = 시작한적있음 and 남은초 <= 0

    # ── 시계 보여 주기 ──
    def _그리기(초, 흐르는중):
        색 = "#c0392b" if 초 <= 0 else ("#e67e22" if 초 <= 60 else "#1a1a1a")
        안내 = "시간이 끝났어요!" if 초 <= 0 else (
            f"남은 시간 ({분}분)" if 흐르는중 else "멈춰 있어요 · ⏱ 시작을 누르세요"
        )
        st.markdown(
            f"""
            <div style="text-align:center; margin:2px 0 6px 0;">
              <div style="font-size:44px; font-weight:800; letter-spacing:2px; color:{색};">
                {_시간글자(초)}
              </div>
              <div style="font-size:13px; color:#888;">{안내}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    if 실행중 and hasattr(st, "fragment"):
        # 1초마다 이 부분만 새로 그려서 시계가 실제로 흘러가게 합니다.
        @st.fragment(run_every=1)
        def _시계():
            초 = _남은초계산(페이지키, 총초)
            _그리기(초, True)
            if 초 <= 0:
                # 시간이 끝나면 화면 전체를 새로 그려 입력칸을 잠급니다.
                try:
                    st.rerun(scope="app")
                except TypeError:      # 예전 버전 대비
                    st.rerun()
        _시계()
    else:
        _그리기(남은초, 실행중)

    if 종료:
        st.error("⏰ 10분이 모두 지났어요! 이제 답을 작성하거나 저장할 수 없습니다. 수고했어요 😊")

    return 남은초, 종료


# ---------------------------------------------------------
# 6) 이름별 저장 공간 만들기
# ---------------------------------------------------------
def _새연구ID():
    """학생마다 자동으로 만들어지는 연구용 번호입니다. (예: R260723-001)"""
    순번 = len(st.session_state["records"]) + 1
    오늘 = datetime.date.today().strftime("%y%m%d")
    return f"R{오늘}-{순번:03d}"


def 이름_저장공간_준비(이름, 학년):
    if 이름 not in st.session_state["records"]:
        st.session_state["records"][이름] = {
            "연구ID": _새연구ID(),   # 자동으로 만들어지는 연구 번호
            "학년": 학년,
            "인지과제": {},   # {문제번호: 학생이 쓴 답}
            "채점": {},       # {문제번호: {"판정": ..., "이유": ...}}  AI 1차 채점 결과
            "글쓰기": [],     # 창의적 글쓰기로 제출한 글들
            "대화": [],       # AI와 나눈 대화 (어느 페이지인지도 함께 기록)
        }
    기록 = st.session_state["records"][이름]
    기록["학년"] = 학년              # 항상 최신 학년으로 갱신
    기록.setdefault("연구ID", _새연구ID())
    기록.setdefault("채점", {})


# ---------------------------------------------------------
# 7) AI 대화 칸을 그려 주는 함수 (두 페이지에서 함께 사용)
#    - 문제_설명을 넣어 주면 AI가 그 문제를 알고 힌트를 줄 수 있어요. (정답은 알려주지 않음)
# ---------------------------------------------------------
def AI대화칸(페이지키, 이름, 학년, 문제_설명=None, 높이=380):
    메시지키 = f"messages_{페이지키}"
    if 메시지키 not in st.session_state:
        st.session_state[메시지키] = []

    # 대화가 길어져도 위쪽 내용이 밀려 사라지지 않도록 높이가 고정된 스크롤 상자를 씁니다.
    대화상자 = st.container(height=높이)
    with 대화상자:
        for 메시지 in st.session_state[메시지키]:
            with st.chat_message(메시지["role"]):
                st.markdown(메시지["content"])

    def 실시간_글자흐름(스트림):
        """AI가 보내주는 조각(chunk)에서 글자만 뽑아 하나씩 내보냅니다."""
        for 조각 in 스트림:
            내용 = 조각.choices[0].delta.content
            if 내용:
                yield 내용

    질문 = st.chat_input("메시지를 입력하세요", key=f"입력_{페이지키}")
    if not 질문:
        return

    if not 이름:
        st.warning("사이드바에서 이름을 먼저 입력해 주세요. 이름으로 대화가 저장됩니다.")
        st.stop()

    # AI에게 줄 안내문: 기본 성격 + (있다면) 지금 풀고 있는 문제 내용
    시스템문 = SYSTEM_PROMPT
    if 문제_설명:
        시스템문 += (
            f"\n\n[학생이 지금 풀고 있는 문제]\n{문제_설명}\n"
            "이 문제의 답은 절대 말하지 말고, 생각하는 방법만 도와줘."
        )

    with 대화상자:
        st.session_state[메시지키].append({"role": "user", "content": 질문})
        with st.chat_message("user"):
            st.markdown(질문)

        with st.chat_message("assistant"):
            try:
                스트림 = 클라이언트().chat.completions.create(
                    model="solar-open2",       # 모델 이름은 반드시 이 글자 그대로!
                    messages=(
                        [{"role": "system", "content": 시스템문}]
                        + st.session_state[메시지키]
                    ),
                    stream=True,               # 스트리밍(글자가 실시간으로 흐름)
                    reasoning_effort="none",   # 추론(생각) 끄기 -> 답이 더 빨리 나옴
                )
                답변 = st.write_stream(실시간_글자흐름(스트림))
            except Exception:
                답변 = None
                st.error(
                    "앗, 답변을 가져오는 데 문제가 생겼어요. 😢\n\n"
                    "잠시 후 다시 시도해 주세요. 문제가 계속되면 인터넷 연결이나 "
                    "API 키(SOLAR_API_KEY) 설정을 확인해 주세요."
                )

    if 답변:
        st.session_state[메시지키].append({"role": "assistant", "content": 답변})
        이름_저장공간_준비(이름, 학년)
        st.session_state["records"][이름]["대화"].append(
            {"페이지": 페이지키, "프롬프트": 질문, "결과물": 답변}
        )
        supabase_db.자동저장(이름)   # 대화 내용을 Supabase에도 저장


# ---------------------------------------------------------
# 8) 현재 작업 초기화 버튼 (저장된 기록은 그대로 남습니다)
#    - 이름칸 / 아이디어칸 / 정답칸 / AI 대화 / 타이머를 모두 비웁니다.
#    - 입력칸의 key 뒤에 붙는 '버전' 번호를 1 올려서, 새 입력칸으로 바꾸는 방식이에요.
# ---------------------------------------------------------
def 모두_초기화(기록도_삭제=False):
    """화면에 있는 것들을 모두 처음 상태로 되돌립니다.

    - 두 페이지의 AI 대화, 타이머(시계), 학생 정보(이름·학년),
      정답 입력칸, 아이디어 입력칸, 문제 번호를 전부 비웁니다.
    - 기록도_삭제=True 이면 저장된 기록(정답·아이디어·대화·채점)까지 완전히 지웁니다.
    """
    # (1) 두 페이지의 AI 대화, 타이머, 자동제출 표시를 모두 비웁니다.
    for 키 in list(st.session_state.keys()):
        이름 = str(키)
        if (이름.startswith("messages_") or 이름.startswith("타이머시작_")
                or 이름.startswith("_자동저장_") or 이름.startswith("_자동제출_")
                or 이름.startswith("_최종제출_") or 이름.startswith("_제출함_")):
            st.session_state.pop(키, None)
    st.session_state["_타이머"] = {}     # 두 페이지 시계를 모두 10:00 으로

    # (2) 학생 정보와 문제 진행 상태를 처음으로
    st.session_state["_이름"] = ""
    st.session_state["_학년"] = "초등학교 5학년"
    st.session_state["문제번호"] = 0
    st.session_state["이전학년"] = None

    # (3) 입력칸(이름·정답·아이디어)을 '새것'으로 바꿔 비웁니다.
    st.session_state["_ver"] = st.session_state.get("_ver", 0) + 1

    # (4) 저장된 기록까지 지울지 여부
    if 기록도_삭제:
        st.session_state["records"] = {}


def 초기화_버튼(페이지키):
    with st.sidebar:
        if st.button("🧹 현재 작업 초기화 (기록은 유지)", key=f"초기화_{페이지키}"):
            모두_초기화(기록도_삭제=False)   # 저장된 기록은 그대로 둡니다.
            st.rerun()


# ---------------------------------------------------------
# 9) AI 1차 채점
#    - 인지과제에 쓴 답을 AI가 먼저 채점해 줍니다.
#    - 결과는 엑셀 파일에서만 볼 수 있고, 학생 화면에는 나오지 않아요.
# ---------------------------------------------------------
def AI_1차채점(이름, 문제은행):
    """한 학생의 인지과제 답을 AI가 채점하고 결과를 저장합니다."""
    정보 = st.session_state["records"].get(이름)
    if not 정보:
        return 0

    문제들 = 문제은행.get(정보.get("학년", ""), [])
    답목록 = 정보.get("인지과제", {})
    if not 답목록:
        return 0

    cli = 클라이언트()
    정보.setdefault("채점", {})
    번호들 = sorted(답목록.keys())
    진행 = st.progress(0.0, text="AI가 채점하는 중이에요...")

    for 순서, 번호 in enumerate(번호들, start=1):
        문제 = 문제들[번호] if 번호 < len(문제들) else {"문제": "", "정답": ""}
        서술형 = str(문제.get("정답", "")).startswith("(정답 없음")

        지시 = (
            "너는 초등학생 인지과제를 채점하는 채점 도우미야. 반드시 한국어로만 답해.\n"
            "아래 형식으로 딱 한 줄만 답해. 다른 말은 절대 쓰지 마.\n"
            "형식:  판정|이유\n"
            "판정은 다음 중 하나만 골라: 정답, 부분정답, 오답, 무응답\n"
            "이유는 30자 이내로 짧게 써.\n"
            + ("이 문제는 정답이 하나가 아닌 서술형이야. 채점 기준을 충족했으면 '정답', "
               "일부만 충족했으면 '부분정답', 거의 못 썼으면 '오답'으로 판정해.\n"
               if 서술형 else "")
            + "학생이 '모르겠다'라고만 썼으면 '무응답'으로 판정해."
        )
        내용 = (
            f"[문제]\n{문제.get('문제', '')}\n\n"
            f"[모범답안 또는 채점 기준]\n{문제.get('정답', '')}\n\n"
            f"[학생이 쓴 답]\n{답목록[번호]}"
        )

        try:
            응답 = cli.chat.completions.create(
                model="solar-open2",       # 모델 이름은 반드시 이 글자 그대로!
                messages=[
                    {"role": "system", "content": 지시},
                    {"role": "user", "content": 내용},
                ],
                reasoning_effort="none",   # 추론(생각) 끄기 -> 빠르게 채점
            )
            글 = (응답.choices[0].message.content or "").strip()
            if "|" in 글:
                판정, 이유 = 글.split("|", 1)
            else:
                판정, 이유 = 글[:6], 글
        except Exception:
            판정, 이유 = "채점 실패", "잠시 후 다시 시도해 주세요."

        정보["채점"][번호] = {"판정": 판정.strip(), "이유": 이유.strip()}
        진행.progress(순서 / len(번호들), text=f"AI가 채점하는 중이에요... ({순서}/{len(번호들)})")

    진행.empty()
    supabase_db.자동저장(이름)   # 채점 결과를 Supabase에도 저장
    return len(번호들)


# ---------------------------------------------------------
# 10) 기록을 보기 좋은 Excel(xlsx)로 만들기
#     - 왼쪽(A~E열)은 인지과제, 오른쪽(G~J열)은 창의적 글쓰기입니다.
#     - 두 과제를 나란히 두어, 대화가 많아도 아래로 길게 내려가지 않아요.
# ---------------------------------------------------------
def _시트이름(원본, 사용중):
    """엑셀 시트 이름에 쓸 수 없는 글자를 지우고, 31자 제한/중복을 처리합니다."""
    금지 = '[]:*?/\\'
    이름 = "".join(c for c in 원본 if c not in 금지).strip() or "학생"
    이름 = 이름[:28]
    후보, n = 이름, 1
    while 후보 in 사용중:
        n += 1
        후보 = f"{이름}_{n}"[:31]
    사용중.add(후보)
    return 후보


def 기록_excel만들기(대상기록, 문제은행):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    wb.remove(wb.active)

    선 = Side(style="thin", color="D9D9D9")
    테두리 = Border(left=선, right=선, top=선, bottom=선)
    가운데 = Alignment(horizontal="center", vertical="center", wrap_text=True)
    왼쪽 = Alignment(horizontal="left", vertical="top", wrap_text=True)
    제목폰트 = Font(bold=True, size=14, color="FFFFFF")
    제목채움 = PatternFill("solid", fgColor="2E7D32")
    구획폰트 = Font(bold=True, size=12, color="1B5E20")
    구획채움 = PatternFill("solid", fgColor="E8F5E9")
    헤더폰트 = Font(bold=True, color="FFFFFF")
    헤더채움A = PatternFill("solid", fgColor="66BB6A")   # 인지과제 쪽
    헤더채움B = PatternFill("solid", fgColor="42A5F5")   # 창의적 글쓰기 쪽
    구획채움B = PatternFill("solid", fgColor="E3F2FD")
    구획폰트B = Font(bold=True, size=12, color="0D47A1")
    라벨폰트 = Font(bold=True)
    노란채움 = PatternFill("solid", fgColor="FFF9C4")    # 연구자가 직접 적는 칸

    사용중 = set()
    for 이름, 정보 in 대상기록.items():
        ws = wb.create_sheet(title=_시트이름(이름, 사용중))

        # 열 너비: A~E = 인지과제 / F = 간격 / G~J = 창의적 글쓰기
        너비 = {"A": 6, "B": 42, "C": 32, "D": 28, "E": 26,
                "F": 3,
                "G": 14, "H": 6, "I": 42, "J": 46}
        for 열, 값 in 너비.items():
            ws.column_dimensions[열].width = 값

        # ── 맨 위: 제목 + 연구ID + 채점 결과란 ──
        ws.merge_cells("A1:E1")
        c = ws["A1"]; c.value = "학생 활동 기록"
        c.font, c.fill, c.alignment = 제목폰트, 제목채움, 가운데
        ws.row_dimensions[1].height = 26

        ws["G1"] = "연구ID"
        ws["G1"].font, ws["G1"].alignment, ws["G1"].border = 라벨폰트, 가운데, 테두리
        ws.merge_cells("H1:J1")
        ws["H1"] = 정보.get("연구ID", "")
        ws["H1"].font = Font(bold=True, size=12)
        ws["H1"].alignment = 왼쪽
        for 열 in ("H", "I", "J"):
            ws[f"{열}1"].border = 테두리

        # 이름 / 학년
        ws["A2"] = "이름"; ws["A2"].font = 라벨폰트; ws["A2"].alignment = 가운데
        ws.merge_cells("B2:C2"); ws["B2"] = 이름; ws["B2"].alignment = 왼쪽
        ws["D2"] = "학년"; ws["D2"].font = 라벨폰트; ws["D2"].alignment = 가운데
        ws["E2"] = 정보.get("학년", ""); ws["E2"].alignment = 왼쪽
        for 셀 in ("A2", "B2", "C2", "D2", "E2"):
            ws[셀].border = 테두리

        # AI 1차 채점 요약 + 연구자가 직접 적는 칸
        채점 = 정보.get("채점", {})
        답개수 = len(정보.get("인지과제", {}))
        세기 = {"정답": 0, "부분정답": 0, "오답": 0, "무응답": 0}
        for v in 채점.values():
            세기[v.get("판정")] = 세기.get(v.get("판정"), 0) + 1
        요약 = (
            f"제출 {답개수}문항 · 정답 {세기.get('정답', 0)} / "
            f"부분정답 {세기.get('부분정답', 0)} / 오답 {세기.get('오답', 0)} / "
            f"무응답 {세기.get('무응답', 0)}"
            if 채점 else "아직 AI 채점을 실행하지 않았습니다."
        )
        ws["G2"] = "AI 1차 채점"; ws["G2"].font, ws["G2"].alignment, ws["G2"].border = 라벨폰트, 가운데, 테두리
        ws.merge_cells("H2:J2"); ws["H2"] = 요약; ws["H2"].alignment = 왼쪽
        for 열 in ("H", "I", "J"):
            ws[f"{열}2"].border = 테두리

        ws["G3"] = "연구자 최종 채점"; ws["G3"].font, ws["G3"].alignment, ws["G3"].border = 라벨폰트, 가운데, 테두리
        ws.merge_cells("H3:J3"); ws["H3"] = ""      # 연구자가 직접 적는 빈칸
        ws["H3"].fill = 노란채움; ws["H3"].alignment = 왼쪽
        for 열 in ("H", "I", "J"):
            ws[f"{열}3"].border = 테두리
            ws[f"{열}3"].fill = 노란채움

        # ─────────── 왼쪽(A~E): 인지과제 ───────────
        r = 5
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(r, 1, "■ 인지과제")
        c.font, c.fill, c.alignment = 구획폰트, 구획채움, 왼쪽
        r += 1
        for i, 글 in enumerate(["번호", "문제", "학생이 쓴 답", "정답(수기 확인용)", "AI 1차 채점"], start=1):
            cc = ws.cell(r, i, 글)
            cc.font, cc.fill, cc.alignment, cc.border = 헤더폰트, 헤더채움A, 가운데, 테두리
        r += 1

        답목록 = 정보.get("인지과제", {})
        문제들 = 문제은행.get(정보.get("학년", ""), [])
        if 답목록:
            for 번호 in sorted(답목록.keys()):
                문제 = 문제들[번호] if 번호 < len(문제들) else {}
                판정정보 = 채점.get(번호, {})
                채점글 = (
                    f"{판정정보.get('판정', '')} · {판정정보.get('이유', '')}"
                    if 판정정보 else "(미채점)"
                )
                ws.cell(r, 1, 번호 + 1).alignment = 가운데
                ws.cell(r, 2, str(문제.get("문제", "")).replace("\n", " ")).alignment = 왼쪽
                ws.cell(r, 3, 답목록[번호]).alignment = 왼쪽
                ws.cell(r, 4, str(문제.get("정답", ""))).alignment = 왼쪽
                ws.cell(r, 5, 채점글).alignment = 왼쪽
                for col in range(1, 6):
                    ws.cell(r, col).border = 테두리
                r += 1
        else:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            ws.cell(r, 1, "(제출한 답이 없습니다)").alignment = 왼쪽
            r += 1

        # 인지과제 페이지에서 나눈 AI 대화도 왼쪽에 이어서
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(r, 1, "■ 인지과제 페이지 · AI 대화")
        c.font, c.fill, c.alignment = 구획폰트, 구획채움, 왼쪽
        r += 1
        ws.cell(r, 1, "번호"); ws.cell(r, 2, "학생 프롬프트")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.cell(r, 3, "AI 답변")
        for col in range(1, 6):
            cc = ws.cell(r, col)
            cc.font, cc.fill, cc.alignment, cc.border = 헤더폰트, 헤더채움A, 가운데, 테두리
        r += 1
        대화_인지 = [d for d in 정보.get("대화", []) if d.get("페이지") == "인지과제"]
        if 대화_인지:
            for i, 쌍 in enumerate(대화_인지, start=1):
                ws.cell(r, 1, i).alignment = 가운데
                ws.cell(r, 2, 쌍.get("프롬프트", "")).alignment = 왼쪽
                ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
                ws.cell(r, 3, 쌍.get("결과물", "")).alignment = 왼쪽
                for col in range(1, 6):
                    ws.cell(r, col).border = 테두리
                r += 1
        else:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            ws.cell(r, 1, "(대화 없음)").alignment = 왼쪽

        # ─────────── 오른쪽(G~J): 창의적 글쓰기 ───────────
        r2 = 5
        ws.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=10)
        c = ws.cell(r2, 7, "■ 창의적 글쓰기 · 제출한 아이디어")
        c.font, c.fill, c.alignment = 구획폰트B, 구획채움B, 왼쪽
        r2 += 1
        ws.cell(r2, 7, "구분"); ws.cell(r2, 8, "번호")
        ws.merge_cells(start_row=r2, start_column=9, end_row=r2, end_column=10)
        ws.cell(r2, 9, "내용")
        for col in range(7, 11):
            cc = ws.cell(r2, col)
            cc.font, cc.fill, cc.alignment, cc.border = 헤더폰트, 헤더채움B, 가운데, 테두리
        r2 += 1

        글목록 = 정보.get("글쓰기", [])
        if 글목록:
            for i, 글 in enumerate(글목록, start=1):
                ws.cell(r2, 7, "아이디어").alignment = 가운데
                ws.cell(r2, 8, i).alignment = 가운데
                ws.merge_cells(start_row=r2, start_column=9, end_row=r2, end_column=10)
                ws.cell(r2, 9, 글).alignment = 왼쪽
                for col in range(7, 11):
                    ws.cell(r2, col).border = 테두리
                r2 += 1
        else:
            ws.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=10)
            ws.cell(r2, 7, "(제출한 아이디어가 없습니다)").alignment = 왼쪽
            r2 += 1

        r2 += 1
        ws.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=10)
        c = ws.cell(r2, 7, "■ 창의적 글쓰기 페이지 · AI 대화")
        c.font, c.fill, c.alignment = 구획폰트B, 구획채움B, 왼쪽
        r2 += 1
        for i, 글 in enumerate(["구분", "번호", "학생 프롬프트", "AI 답변"], start=7):
            cc = ws.cell(r2, i, 글)
            cc.font, cc.fill, cc.alignment, cc.border = 헤더폰트, 헤더채움B, 가운데, 테두리
        r2 += 1
        대화_글 = [d for d in 정보.get("대화", []) if d.get("페이지") != "인지과제"]
        if 대화_글:
            for i, 쌍 in enumerate(대화_글, start=1):
                ws.cell(r2, 7, "대화").alignment = 가운데
                ws.cell(r2, 8, i).alignment = 가운데
                ws.cell(r2, 9, 쌍.get("프롬프트", "")).alignment = 왼쪽
                ws.cell(r2, 10, 쌍.get("결과물", "")).alignment = 왼쪽
                for col in range(7, 11):
                    ws.cell(r2, col).border = 테두리
                r2 += 1
        else:
            ws.merge_cells(start_row=r2, start_column=7, end_row=r2, end_column=10)
            ws.cell(r2, 7, "(대화 없음)").alignment = 왼쪽

    저장통 = io.BytesIO()
    wb.save(저장통)
    return 저장통.getvalue()


# ---------------------------------------------------------
# 10-2) 연구자 패널 (사이드바)
#       - 암호 '0000'을 맞게 입력해야 채점·내려받기·삭제를 할 수 있습니다.
# ---------------------------------------------------------
def 연구자_패널(페이지키, 문제은행):
    with st.sidebar:
        st.divider()
        st.subheader("저장된 기록 (연구자용)")

        암호 = st.text_input(
            "기록 확인 암호", type="password",
            placeholder="연구자만 입력", key=f"암호_{페이지키}",
        )
        암호맞음 = (암호 == "0000")

        if not 암호맞음:
            st.caption("🔒 기록을 보려면 암호를 바르게 입력하세요.")
            return False

        if not st.session_state["records"]:
            st.caption("아직 저장된 기록이 없어요.")
            return True

        이름목록 = ["(전체)"] + list(st.session_state["records"].keys())
        선택 = st.selectbox("학생 선택", 이름목록, key=f"선택_{페이지키}")

        if 선택 == "(전체)":
            대상이름들 = list(st.session_state["records"].keys())
        else:
            대상이름들 = [선택]

        # AI 1차 채점 실행 버튼 (누를 때만 채점하므로 시간이 조금 걸려요)
        if st.button("🤖 AI 1차 채점 실행", key=f"채점_{페이지키}"):
            총합 = 0
            for 학생 in 대상이름들:
                총합 += AI_1차채점(학생, 문제은행)
            if 총합:
                st.success(f"채점 완료! ({총합}문항) 엑셀에서 확인하세요.")
            else:
                st.info("채점할 답이 없어요.")

        # ── Supabase(온라인 데이터베이스) 저장 ──
        st.divider()
        st.caption("☁️ Supabase 저장")

        if not supabase_db.사용가능():
            st.warning(
                "Supabase 접속 정보가 없어요.\n\n"
                "Secrets에 SUPABASE_URL 과 SUPABASE_KEY 를 넣어 주세요."
            )
        else:
            # 지금 넣은 키가 올바른 종류인지 알려 줍니다.
            종류, 안내 = supabase_db.키종류()
            if 종류 == "service_role":
                st.caption(f"🔑 {안내}")
            else:
                st.error(
                    f"🔑 {안내}\n\n"
                    "Supabase → Project Settings → API 에서 "
                    "**service_role(secret)** 키를 복사해 Secrets의 "
                    "SUPABASE_KEY 에 넣어 주세요."
                )

            if st.button("☁️ Supabase에 저장하기", key=f"업로드_{페이지키}"):
                성공수, 실패목록 = 0, []
                for 학생 in 대상이름들:
                    성공, 메시지 = supabase_db.학생저장(학생, 조용히=True)
                    if 성공:
                        성공수 += 1
                    else:
                        실패목록.append(f"{학생}: {메시지}")
                if 성공수:
                    st.success(f"{성공수}명의 기록을 저장했어요.")
                for 줄 in 실패목록:
                    st.error(줄)

            # 지금 Supabase에 몇 명이 저장되어 있는지 확인
            if st.button("🔎 저장 현황 확인", key=f"현황_{페이지키}"):
                수, 오류 = supabase_db.저장현황()
                if 오류:
                    st.error(f"확인 실패 · {오류}")
                else:
                    st.info(f"Supabase에 저장된 참가자: {수}명")

            # 학생이 답을 저장할 때마다 자동으로 올라간 결과
            최근 = st.session_state.get("_supabase_최근")
            if 최근:
                st.caption(f"최근 자동 저장 · {최근}")

        st.divider()
        st.caption(
            "⚠️ 아래 버튼은 **모든 것을 처음 상태로** 되돌립니다.\n\n"
            "저장된 기록(정답·아이디어·AI 대화·채점 결과)은 물론,\n"
            "학생 정보·입력칸·시계까지 전부 지워지며 되돌릴 수 없습니다."
        )
        동의 = st.checkbox("삭제에 동의합니다", key=f"동의_{페이지키}")
        if st.button("🗑 전체 초기화 (기록까지 모두 삭제)", disabled=not 동의, key=f"삭제_{페이지키}"):
            모두_초기화(기록도_삭제=True)   # 기록을 포함해 전부 초기화
            st.rerun()

        return True


# ---------------------------------------------------------
# 11) 페이지 머리글 (두 페이지가 똑같은 모양으로 보이게 하는 함수)
#     - 맨 위에 오늘 날짜가 자동으로 들어갑니다.
#     - 아이콘 + 제목 + 설명 순서로 그려집니다.
# ---------------------------------------------------------
def 페이지_머리글(아이콘, 제목, 설명):
    오늘 = datetime.date.today().strftime("%Y년 %m월 %d일")
    st.caption(f"📅 {오늘}")          # 오늘 날짜 (자동)
    st.title(f"{아이콘} {제목}")       # 큰 제목
    st.caption(설명)                  # 제목 아래 설명 한 줄
